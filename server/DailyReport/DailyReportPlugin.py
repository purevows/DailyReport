from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, colors
from functools import wraps
from flask import session, redirect, jsonify
import datetime
import os
import logging


def authority_required(authority="user"):
    def wrapper(func):
        @wraps(func)
        def decorated_view(*args, **kwargs):
            if authority != "user":
                if authority == "综合管理部" and session.get("department") != "综合管理部":
                    return jsonify({"status": 401})
            return func(*args, **kwargs)
        return decorated_view
    return wrapper


def login_required(func):
    @wraps(func)
    def decorated_view(*args, **kwargs):
        if session.get("token") is None:
            return redirect("/login")
        return func(*args, **kwargs)
    return decorated_view


def log():
    log_mgr = logging.getLogger(__name__)
    log_mgr.setLevel(logging.INFO)
    if not log_mgr.handlers:
        file_handler = logging.FileHandler("../serverlog/DailyReport.log", encoding="utf-8")
        formatter = logging.Formatter(fmt="%(asctime)s %(levelname)s %(filename)s %(message)s",
                                      datefmt="%Y/%m/%d %X")
        file_handler.setFormatter(formatter)
        log_mgr.addHandler(file_handler)
    return log_mgr


logger = log()


MessageSetting = {
    "target_url": "http://172.23.5.47/sms/mt",
    "cid": "0041",
    "pass": "1234",
    "content": "（开发区交投）亲记得填报今日事项哦~"
}


def write_to_workbook(ret, user, timestamp):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = timestamp.replace("-", ".")
    worksheet.merge_cells("A1:E1")
    worksheet.merge_cells("A2:E2")
    worksheet.column_dimensions["A"].width = 6
    worksheet.column_dimensions["B"].width = 15
    worksheet.column_dimensions["C"].width = 46
    worksheet.column_dimensions["D"].width = 13
    worksheet.column_dimensions["E"].width = 10
    worksheet.cell(1, 1).value = "集团今日事项落实完成情况表"
    worksheet.cell(1, 1).font = FontSettings["Title"]
    worksheet.cell(1, 1).alignment = AlignmentSettings["Title"]
    # worksheet.cell(1, 1).border = BorderSettings["Public"]
    worksheet.cell(2, 1).value = "日期 " + timestamp.replace("-", ".")
    worksheet.cell(2, 1).font = FontSettings["Title"]
    worksheet.cell(2, 1).alignment = AlignmentSettings["DateTime"]
    # worksheet.cell(2, 1).border = BorderSettings["Public"]
    worksheet["A3"] = "序号"
    worksheet["A3"].font = FontSettings["Title"]
    worksheet["A3"].alignment = AlignmentSettings["Title"]
    worksheet["A3"].border = BorderSettings["Public"]
    worksheet["B3"] = "负责部门"
    worksheet["B3"].font = FontSettings["Title"]
    worksheet["B3"].alignment = AlignmentSettings["Title"]
    worksheet["B3"].border = BorderSettings["Public"]
    worksheet["C3"] = "完成情况"
    worksheet["C3"].font = FontSettings["Title"]
    worksheet["C3"].alignment = AlignmentSettings["Title"]
    worksheet["C3"].border = BorderSettings["Public"]
    worksheet["D3"] = "经办人"
    worksheet["D3"].font = FontSettings["Title"]
    worksheet["D3"].alignment = AlignmentSettings["Title"]
    worksheet["D3"].border = BorderSettings["Public"]
    worksheet["E3"] = "备注"
    worksheet["E3"].font = FontSettings["Title"]
    worksheet["E3"].alignment = AlignmentSettings["Title"]
    worksheet["E3"].border = BorderSettings["Public"]
    next_line = 4
    num = 1
    j = 0
    for key, value in department_sequence.items():
        for i in range(j, len(ret)):
            if ret[i]["department"] == key:
                temp = ret[i]
                ret[i] = ret[j]
                ret[j] = temp
                j = j + 1
    j = 0
    while j < len(ret):
        for i in range(j + 1, len(ret)):
            if ret[j]["operator"] == ret[i]["operator"] and ret[j]["department"] == ret[i]["department"]:
                temp = ret[i]
                ret[i] = ret[j + 1]
                ret[j + 1] = temp
                j = j + 1
        j = j + 1
    for data in ret:
        # if data["department"] == key:
        worksheet["A" + str(next_line)] = num
        worksheet["A" + str(next_line)].font = FontSettings["Body_Num"]
        worksheet["A" + str(next_line)].alignment = AlignmentSettings["Body"]
        worksheet["A" + str(next_line)].border = BorderSettings["Public"]
        num = num + 1
        worksheet["B" + str(next_line)] = department_sequence[data["department"]]
        worksheet["B" + str(next_line)].font = FontSettings["Body_Department"]
        worksheet["B" + str(next_line)].alignment = AlignmentSettings["Body"]
        worksheet["B" + str(next_line)].border = BorderSettings["Public"]
        worksheet["C" + str(next_line)] = data["content"].replace("<br>", "\n")
        worksheet["C" + str(next_line)].font = FontSettings["Body"]
        worksheet["C" + str(next_line)].alignment = AlignmentSettings["Body_Content"]
        worksheet["C" + str(next_line)].border = BorderSettings["Public"]
        worksheet["D" + str(next_line)] = data["operator"].replace("<br>", "\n")
        worksheet["D" + str(next_line)].font = FontSettings["Body"]
        worksheet["D" + str(next_line)].alignment = AlignmentSettings["Body"]
        worksheet["D" + str(next_line)].border = BorderSettings["Public"]
        worksheet["E" + str(next_line)] = data["remark"].replace("<br>", "\n")
        worksheet["E" + str(next_line)].font = FontSettings["Body"]
        worksheet["E" + str(next_line)].alignment = AlignmentSettings["Body"]
        worksheet["E" + str(next_line)].border = BorderSettings["Public"]
        next_line = next_line + 1
    worksheet.merge_cells("A" + str(next_line) + ":E" + str(next_line))
    worksheet.cell(next_line, 1).value = "导出时间：" + datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") +\
                                         " " + user
    worksheet.cell(next_line, 1).font = FontSettings["Body"]
    worksheet.cell(next_line, 1).alignment = AlignmentSettings["Body_Content"]
    for i in range(1, next_line):
        if i >= 4:
            c_rows = worksheet["C" + str(i)].value.count("\n")
            d_rows = worksheet["D" + str(i)].value.count("\n")
            e_rows = worksheet["E" + str(i)].value.count("\n")
            max_row = max(c_rows, d_rows, e_rows)
            if max_row > 1:
                worksheet.row_dimensions[i].height = 30 + 15 * (max_row - 1)
            else:
                worksheet.row_dimensions[i].height = 30
        else:
            worksheet.row_dimensions[i].height = 30
    i = 4
    while i < next_line:
        if worksheet["D" + str(i)].value == worksheet["D" + str(i + 1)].value:
            j = i
            while worksheet["D" + str(j)].value == worksheet["D" + str(j + 1)].value:
                j = j + 1
            worksheet.merge_cells("D" + str(i) + ":D" + str(j))
            i = j
        i = i + 1
    i = 4
    while i < next_line:
        if worksheet["B" + str(i)].value == worksheet["B" + str(i + 1)].value:
            j = i
            while worksheet["B" + str(j)].value == worksheet["B" + str(j + 1)].value:
                j = j + 1
            worksheet.merge_cells("B" + str(i) + ":B" + str(j))
            i = j
        i = i + 1
    file_name = "今日事项落实完成情况（" + timestamp.replace("-", ".") + ").xlsx"
    workbook.save(os.getcwd() + r"/DailyReport/ExportFile/" + file_name)
    logger.info("Items are exported by " + user + ", filename:" + file_name)
    # workbook.save(os.getcwd() + "/" + file_name)
    return file_name


department_sequence = {
    "规划与建设管理部（建设公司）": "规建部（建设公司）",
    "资产运营部（置业公司）": "资产部（置业公司）",
    "战略与投资发展部": "战略部",
    "党群工作部": "党群部",
    "人力资源部": "人事部",
    "综合管理部": "综合部",
    "开泰公司": "开泰公司",
    "站场公司": "站场公司"
}
FontSettings = {
    "Title": Font(name=u"宋体", bold=True, size=11),
    "Body_Num": Font(name=u"宋体", size=11),
    "Body": Font(name=u"宋体", size=10),
    "Body_Department": Font(name=u"宋体", bold=True, size=10)
}
BorderSettings = {
    "Public": Border(left=Side(style="thin", color=colors.BLACK),
                     right=Side(style="thin", color=colors.BLACK),
                     top=Side(style="thin", color=colors.BLACK),
                     bottom=Side(style="thin", color=colors.BLACK))
}
AlignmentSettings = {
    "Title": Alignment(horizontal="center", vertical="center", wrap_text=True),
    "DateTime": Alignment(horizontal="left", vertical="center", wrap_text=True),
    "Body": Alignment(horizontal="center", vertical="center", wrap_text=True),
    "Body_Content": Alignment(horizontal="left", vertical="center", wrap_text=True)
}
'''
ret = [{'id': '375045c488ef11eabb910800274f6112', 'content': '555555', 'department': '综合管理部',
        'operator': 'yyyyyyyy', 'remark': 'hhhhhhhh', 'uploader': '鲁剑涛',
        'upload_time': datetime.datetime(2020, 4, 28, 9, 25, 59), 'first_check': '谭宗琦',
        'first_check_time': datetime.datetime(2020, 4, 28, 11, 17, 12), 'double_check': '戴涛',
        'double_check_time': datetime.datetime(2020, 4, 28, 11, 18, 3)},
       {'id': '5e4c756088ec11ea95c60800274f6112', 'content': 'gggggg', 'department': '战略与投资发展部',
        'operator': 'dddddd', 'remark': 'dhhhhhh', 'uploader': '梁伟健',
        'upload_time': datetime.datetime(2020, 4, 28, 9, 5, 36), 'first_check': '谭宗琦',
        'first_check_time': datetime.datetime(2020, 4, 28, 11, 17, 12), 'double_check': '戴涛',
        'double_check_time': datetime.datetime(2020, 4, 28, 11, 18, 3)},
       {'id': '8e9bf56088f611ea99320800274f6112', 'content': 'rrrr', 'department': '战略与投资发展部',
        'operator': 'tttttt', 'remark': 'tyyy', 'uploader': '郭建万',
        'upload_time': datetime.datetime(2020, 4, 28, 10, 18, 32), 'first_check': '谭宗琦',
        'first_check_time': datetime.datetime(2020, 4, 28, 11, 17, 13), 'double_check': '戴涛',
        'double_check_time': datetime.datetime(2020, 4, 28, 11, 18, 3)},
       {'id': '97c8abf688f611ea82460800274f6112', 'content': 'ddd', 'department': '战略与投资发展部',
        'operator': 'gggggg', 'remark': 'hgggg', 'uploader': '郭建万',
        'upload_time': datetime.datetime(2020, 4, 28, 10, 18, 47), 'first_check': '谭宗琦',
        'first_check_time': datetime.datetime(2020, 4, 28, 11, 17, 13), 'double_check': '戴涛',
        'double_check_time': datetime.datetime(2020, 4, 28, 11, 18, 3)},
       {'id': '97cbe7d088f611ea82460800274f6112', 'content': '44', 'department': '战略与投资发展部',
        'operator': '666', 'remark': '000', 'uploader': '郭建万',
        'upload_time': datetime.datetime(2020, 4, 28, 10, 18, 47), 'first_check': '谭宗琦',
        'first_check_time': datetime.datetime(2020, 4, 28, 11, 17, 13), 'double_check': '戴涛',
        'double_check_time': datetime.datetime(2020, 4, 28, 11, 18, 3)},
       {'id': '97d36d2088f611ea82460800274f6112', 'content': '555', 'department': '战略与投资发展部',
        'operator': '7777', 'remark': '888', 'uploader': '郭建万',
        'upload_time': datetime.datetime(2020, 4, 28, 10, 18, 47), 'first_check': '谭宗琦',
        'first_check_time': datetime.datetime(2020, 4, 28, 11, 17, 13), 'double_check': '戴涛',
        'double_check_time': datetime.datetime(2020, 4, 28, 11, 18, 3)},
       {'id': 'cce3a1de88fa11eab69c0800274f6112', 'content': 'fff', 'department': '综合管理部',
        'operator': 'ggg', 'remark': 'qqq', 'uploader': '鲁剑涛',
        'upload_time': datetime.datetime(2020, 4, 28, 10, 48, 54), 'first_check': '谭宗琦',
        'first_check_time': datetime.datetime(2020, 4, 28, 11, 17, 13), 'double_check': '戴涛',
        'double_check_time': datetime.datetime(2020, 4, 28, 11, 18, 3)},
       {'id': 'cce5a88088fa11eab69c0800274f6112', 'content': 'ddd', 'department': '综合管理部',
        'operator': 'hhh', 'remark': 'www', 'uploader': '鲁剑涛',
        'upload_time': datetime.datetime(2020, 4, 28, 10, 48, 54), 'first_check': '谭宗琦',
        'first_check_time': datetime.datetime(2020, 4, 28, 11, 17, 13), 'double_check': '戴涛',
        'double_check_time': datetime.datetime(2020, 4, 28, 11, 18, 3)},
       {'id': 'ccec8f0688fa11eab69c0800274f6112', 'content': '333', 'department': '综合管理部',
        'operator': 'jjj', 'remark': 'fff', 'uploader': '鲁剑涛',
        'upload_time': datetime.datetime(2020, 4, 28, 10, 48, 54), 'first_check': '谭宗琦',
        'first_check_time': datetime.datetime(2020, 4, 28, 11, 17, 13), 'double_check': '戴涛',
        'double_check_time': datetime.datetime(2020, 4, 28, 11, 18, 3)},
       {'id': 'cceede1e88fa11eab69c0800274f6112', 'content': '455', 'department': '综合管理部',
        'operator': '333', 'remark': 'ggg', 'uploader': '鲁剑涛',
        'upload_time': datetime.datetime(2020, 4, 28, 10, 48, 54), 'first_check': '谭宗琦',
        'first_check_time': datetime.datetime(2020, 4, 28, 11, 17, 13), 'double_check': '戴涛',
        'double_check_time': datetime.datetime(2020, 4, 28, 11, 18, 3)},
       {'id': 'ccf1a3ce88fa11eab69c0800274f6112', 'content': '666', 'department': '综合管理部',
        'operator': '555', 'remark': 'hhh', 'uploader': '鲁剑涛',
        'upload_time': datetime.datetime(2020, 4, 28, 10, 48, 54), 'first_check': '谭宗琦',
        'first_check_time': datetime.datetime(2020, 4, 28, 11, 17, 13), 'double_check': '戴涛',
        'double_check_time': datetime.datetime(2020, 4, 28, 11, 18, 3)},
       {'id': 'ccf54ede88fa11eab69c0800274f6112', 'content': '777', 'department': '综合管理部',
        'operator': '666', 'remark': 'jjj', 'uploader': '鲁剑涛',
        'upload_time': datetime.datetime(2020, 4, 28, 10, 48, 55), 'first_check': '谭宗琦',
        'first_check_time': datetime.datetime(2020, 4, 28, 11, 17, 13), 'double_check': '戴涛',
        'double_check_time': datetime.datetime(2020, 4, 28, 11, 18, 3)}]


if __name__ == "__main__":
    write_to_workbook(ret, "张惠")
'''
